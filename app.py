# -*- coding: utf-8 -*-
"""
Created on Fri Oct 26 16:16:09 2018

@author: xutao
"""


#加载数据库就要带着这个了
import os
import sys

import json
import ast
import pickle
import winreg 

#自动创建数据库加载的
import click



from flask_wtf import FlaskForm #wtf运用
from wtforms import StringField,SelectField,SelectMultipleField #选择
from flask import Flask,request,render_template,redirect,url_for,jsonify, make_response
import traceback
import openpyxl
from flask_bootstrap import Bootstrap #渲染




from flask import Flask, render_template, request, url_for, redirect, flash
from flask_sqlalchemy import SQLAlchemy  # 导入数据库扩展类


wb_path='excel.xlsx'
wb=openpyxl.load_workbook(wb_path)        
sheet=wb['Sheet1']     
maxrow=sheet.max_row  #读取数据


#对不同系统进行判断
WIN = sys.platform.startswith('win')
if WIN:  # 如果是 Windows 系统，使用三个斜线
    prefix = 'sqlite:///'
else:  # 否则使用四个斜线
    prefix = 'sqlite:////'


app = Flask(__name__)


#Flask 提供了一个统一的接口来写入和获取这些配置变量：Flask.config 字典
#配置变量的名称必须使用大写，写入配置的语句一般会放到扩展类实例化语句之前
#app.root_path 返回程序实例所在模块的路径
app.config['SQLALCHEMY_DATABASE_URI'] = prefix + os.path.join(app.root_path, 'data.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False  # 关闭对模型修改的监控
app.config['SECRET_KEY'] = 'dev'  # 等同于 app.secret_key = 'dev'
app.config['JSON_AS_ASCII'] = False


# 在扩展类实例化前加载配置
db = SQLAlchemy(app)  # 初始化扩展，传入程序实例 app
bootstrap = Bootstrap(app)


#自定义命令来自动执行创建数据库表操作
@app.cli.command()  # 注册为命令
@click.option('--drop', is_flag=True, help='Create after drop.')  # 设置选项
def initdb(drop):
    """Initialize the database."""
    if drop:  # 判断是否输入了选项
        db.drop_all()
    db.create_all()
    click.echo('Initialized database.')  # 输出提示信息


#模型类要声明继承 db.Model
class Parameter(db.Model):  # 表名将会是
    id = db.Column(db.Integer, primary_key=True)  # 主键
    name = db.Column(db.String(60))  
    value = db.Column(db.String(60)) 

class Form(FlaskForm):

    #stage_user_defined_feature_columns_name1 = StringField('', validators=[]) #输入数据
    s1=SelectField('stage_backend',choices=[])
    s2=SelectField('stage_package',choices=[])
    s3=SelectField('stage_name',choices=[])
    s4=SelectMultipleField('stage_param',choices=[])

    stage_user_defined_feature_columns_name2 = StringField('stage_user_defined_feature_columns_name:', validators=[]) #输入数据    
    s5=SelectField('stage_backend',choices=[])
    s6=SelectField('stage_package',choices=[])
    s7=SelectField('stage_name',choices=[])
    s8=SelectMultipleField('stage_param',choices=[]) 
	
    stage_user_defined_feature_columns_name3 = StringField('stage_user_defined_feature_columns_name:', validators=[]) #输入数据
    s9=SelectField('stage_backend',choices=[])
    s10=SelectField('stage_package',choices=[])
    s11=SelectField('stage_name',choices=[])
    s12=SelectMultipleField('stage_param',choices=[])   

    stage_user_defined_feature_columns_name4 = StringField('stage_user_defined_feature_columns_name:', validators=[]) #输入数据
    s13=SelectField('stage_backend',choices=[])
    s14=SelectField('stage_package',choices=[])
    s15=SelectField('stage_name',choices=[])
    s16=SelectMultipleField('stage_param',choices=[]) 

result=[] #结果列表


#自定义命令函数把虚拟数据添加到数据库里
@app.cli.command()
def forge():
    """Generate fake data."""
    db.create_all()

    # 全局的两个变量移动到这个函数内

    datas = [
	    {'name': 'pipeline_uid', 'value': '2019'},
        {'name': 'root_directory', 'value': './sklearn_search_ranking_root_directory/'},
        {'name': 'root_data_directory', 'value': './sklearn_search_ranking_root_directory/data'},
        {'name': 'pipeline_input_format', 'value': 'file'},
        {'name': 'pipeline_input_file_format', 'value': 'csv'},
        {'name': 'pipeline_input_feature_columns_name', 'value': '[\'f1\', \'f2\', \'f3\', \'f4\']'},
        {'name': 'pipeline_input_label_columns_name', 'value': '[\'label\']'},
        {'name': 'pipeline_stages_config_dict', 'value': 'pipeline_stages_config_dict'},
        {'name': 'sampleID', 'value': 'sampleID'},
        {'name': 'groupID', 'value': 'groupID'},
        {'name': 'pipeline_input_file_name', 'value': 'data_eval'},
        {'name': 'pipeline_phase_config', 'value': 'train'},
    ]

    for da in datas:
        parameter = Parameter(name=da['name'], value=da['value'])
        db.session.add(parameter)

    db.session.commit()
    click.echo('Done.')



@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':  # 判断是否是 POST 请求
		# Flask 会在请求触发后把请求信息放到 request 对象里
        # 获取表单数据
        name = request.form.get('name')  # 传入表单对应输入字段的 name 值
        value = request.form.get('value')
        # 验证数据
        if not name or not value:
            flash('Invalid input.')  # 显示错误提示
            return redirect(url_for('index'))  # 重定向回主页
        # 保存表单数据到数据库
        parameter = Parameter(name=name, value=value)  # 创建记录
        db.session.add(parameter)  # 添加到数据库会话
        db.session.commit()  # 提交数据库会话
        flash('Item created.')  # 显示成功创建的提示

		#Flask 提供了 redirect() 函数来快捷生成这种重定向响应，传入重定向的目标 URL 
        return redirect(url_for('index'))  # 重定向回主页

    parameters = Parameter.query.all()
    return render_template('index.html', parameters=parameters)



@app.errorhandler(404)  # 传入要处理的错误代码
def page_not_found(e):  # 接受异常对象作为参数
    #user = User.query.first()
    return render_template('404.html'), 404  # 返回模板和状态码


#编辑电影条目
#url
@app.route('/parameter/edit/<int:parameter_id>', methods=['GET', 'POST'])
def edit(parameter_id): #执行函数 传入数据库id
	#get_or_404()方法，它会返回对应主键的记录，如果没有找到，则返回 404 错误响应
    parameter = Parameter.query.get_or_404(parameter_id) #获取该id数据

    if request.method == 'POST':  # 处理编辑表单的提交请求
        name = request.form['name']
        value = request.form['value']
        
		#判断输入的数据是否有问题
        if not name or not value:
            flash('Invalid input.')
            return redirect(url_for('edit', parameter_id=parameter_id))  # 重定向回对应的编辑页面

        parameter.name = name  # 更新标题
        parameter.value = value  # 更新年份
        db.session.commit()  # 提交数据库会话
        flash('Item updated.') #flash提示信息
        return redirect(url_for('index'))  # 重定向回主页

    return render_template('edit.html', parameter=parameter)  # 传入被编辑的电影记录



@app.route('/parameter/delete/<int:parameter_id>', methods=['POST'])  # 限定只接受 POST 请求
def delete(parameter_id):
    parameter = Parameter.query.get_or_404(parameter_id)  # 获取电影记录
    db.session.delete(parameter)  # 删除对应的记录
    db.session.commit()  # 提交数据库会话
    flash('Item deleted.')
    return redirect(url_for('index'))  # 重定向回主页


# 选择标志
select_list = ['s1', 'stage_user_defined_feature_columns_name2', 'stage_user_defined_feature_columns_name3', 'stage_user_defined_feature_columns_name4']

# 获取数据库数据
def DB_data():
    #parameter = Parameter.query.first()
    #data = {}
    #data[str(parameter.name)] = parameter.value

    # 新建stage参数字典
    stage_param_dict = {}
    # 数据库参数数据
    num = Parameter.query.count()
    
    i = 0 #循环计数
    while i < num:
        i +=1
        data = Parameter.query.get(i)
        stage_param_dict[str(data.name)] = data.value

    return stage_param_dict

# 处理数据形式函数
def test(data_dict):
    # 返回当前stage数据条目化并列表化
    stage_data_list = list(data_dict.items())

    #stage_data_len = len(stage_data_list)
    result = {}  # 设结果字典
    
    # 通过是否有stage_user_defined_feature_columns_name 区分处理方式
    if stage_data_list[2][0][:-1] != 'stage_user_defined_feature_columns_name':
        #result[stage_data_list[1][0]] = stage_data_list[1][1]
        
        # 向结果字典添加stage_backend、stage_name两项参数
        result['stage_backend'] = stage_data_list[2][1]
        stage_name = stage_data_list[4][1].split('-')[2]
        result['stage_name'] = stage_name
        
        # 新建参数字典
        stage_param_dict = {}
        # 对参数相开始循环处理  后期可以用循环判定的方式 不然位数不见得正确
        for data in stage_data_list[6:]:
            stage_param_dict[data[0].split('-')[4]] = data[1]
        
        # 然后将参数字典放入结果字典的
        result[stage_data_list[0][0]] = stage_data_list[0][1]
        result[stage_data_list[1][0]] = stage_param_dict
    
    # 如果不是首stage的话
    else:
        # 向结果字典添加stage_backend、stage_name两项参数
        result['stage_backend'] = stage_data_list[3][1]
        stage_name = stage_data_list[5][1].split('-')[2]
        result['stage_name'] = stage_name
    
        # 新建参数字典
        stage_param_dict = {}
    
        # 对参数数据行循环处理
        for data in stage_data_list[7:]:
            #print(data[0].split('-'))
            stage_param_dict[data[0].split('-')[4]] = data[1]
    
        # 然后将参数字典放入结果字典的
        result[stage_data_list[0][0]] = stage_data_list[0][1]
        result[stage_data_list[1][0]] = stage_param_dict
    
        # stage_user_defined_feature_columns_name 数据添加
        stage_user_defined_feature_columns_name = stage_data_list[2][0][:-1]
        stage_user_defined_feature_columns_name_list = list((stage_data_list[2][1]).split(' '))
        name_list = [] # 循环处理name
        for columns_name in stage_user_defined_feature_columns_name_list:
            if columns_name == '':
                continue
            else:
                name_list.append(columns_name)
        result[stage_user_defined_feature_columns_name] = name_list
        
    return result

                         
# 获取数据函数json
def SE_data(test_data):
    data_dict = test_data # 获取测试数据字典
    data_len = len(data_dict) # 获取数据字典长度
    # 数据key值列表  value值列表
    data_key = []
    data_value = []
    for data in data_dict:
        data_key.append(data)
        data_value.append(data_dict[data])
    # 获取测试数据条目列表
    data_item_list = list(data_dict.items())
    # 所需设置
    stage_data_dict = {} # stage总数据字典
    result_list = [] # 
    #stage_name_dict = {'key':'stage1_config_dict'}
    stage_name_dict = {} # stage名字字典
    stage_name_str = '' # stage_name比较储存
    #pipeline_stages_config_dict = [] # pipeline_stages_config_dict参数预留
    stage = 1 # stage计数
    # 按数据长度循环所有数据
    for i in range(data_len): 

        # 获取当前i的key 和 value两个值
        key = data_key[i]
        value = data_value[i]
    
        # 对key进行判定，是否为stage开头标志
        if key in select_list:
            if key == 's1': # 如果是第一个
        
                # 当前i为stage数据首项的取值 往后挪四位 可以获取stage_name取值
                e = i + 4
                if len(data_item_list[e][0].split('-')) > 1:
                    data_item = data_item_list[e][0].split('-')[3] # 取到stage_name的值
                    stage_name_str = data_item # 储存stage_name字符串

                    stage_name = 'stage1_config_dict' # 第一个stage数据字典  手动定义 用于比较
                    stage_dict = 'stage1_config_dict' # 用于生成字典名
                    stage_dict = {} # 新建数据字典
            
                    # stage_param_dict的数据
                    stage_param_dict = 'stage_param_dict' + str(int(stage))
                    #pipeline_stages_config_dict.append(stage_param_dict)  #另一种直接的形式
                    stage_dict['stage_param_dict'] = stage_param_dict
            
                    # 第一个stage特殊处理 在此添加参数  stage名字字典  stage1_config_dict
                    stage_name_dict['key'] = stage_name 
            
                    # 具体参数字典提前占位
                    stage_dict[str(stage_param_dict)] = '' 

                # 否则 如果没有参数 则
                else:
                    e = i + 2
                    data_item = data_item_list[e][1].split('-')[2]
                    stage_name_str = data_item # 储存stage_name字符串

                    stage_name = 'stage1_config_dict' # 第一个stage数据字典  手动定义 用于比较
                    stage_dict = 'stage1_config_dict' # 用于生成字典名
                    stage_dict = {} # 新建数据字典        
        
                    # stage_param_dict的数据
                    stage_param_dict = 'stage_param_dict' + str(int(stage))
                    #pipeline_stages_config_dict.append(stage_param_dict)  #另一种直接的形式
                    stage_dict['stage_param_dict'] = stage_param_dict
 
                    # 第一个stage特殊处理 在此添加参数  stage名字字典  stage1_config_dict
                    stage_name_dict['key'] = stage_name 
            
                    # 具体参数字典提前占位
                    stage_dict[str(stage_param_dict)] = '' 
      
            # 否则不是第一个stage
            else:
                stage += 1 # stage自增
            
                # 如果有stage_user_defined_feature_columns_name
                # 此时具体名为当前的 s

                #形成当前的字典名
                stage_name = 'stage' + str(int(stage)) + '_config_dict'
                stage_dict = 'stage' + str(int(stage)) + '_config_dict'
                stage_dict = {} #进行新建
            
                # stage_param_dict的数据
                stage_param_dict = 'stage_param_dict' + str(int(stage))
                #pipeline_stages_config_dict.append(stage_param_dict)
                stage_dict['stage_param_dict'] = stage_param_dict
            
                # 具体参数字典提前占位
                stage_dict[str(stage_param_dict)] = ''   
    
        # 添加每条数据
        if key != 's1' and key in select_list: # 如果新换stage 生成前台数据数据与例行数据组合
            if stage_name_str != '':
                stage_dict[key] = value + ' stage' + str(int(stage) - 1) + '_' + str(stage_name_str)
        
                # 然后更新data_item

                e = i + 3
                if str(data_item_list[i+1][0])[:-1] == 'stage_user_defined_feature_columns_name':
                    data_item = ''

                else:
                    if len(data_item_list[e][0].split('-')) > 1:
                        data_item = data_item_list[e][0].split('-')[3]
                    else:
                        e = i + 3
                        data_item = data_item_list[e][1].split('-')[2]
            
                stage_name_str = data_item # 储存stage_name列表
            else:
                continue
        else:
            stage_dict[key] = value # 否则正常添加数据

        # 数据添加result_list
        result_list.append(stage_dict)
    
        if stage_name_dict['key'] != stage_name or i == data_len - 1: # 如果不相等 证明换stage了
            i -= 1 # 获取上一个stage最后完整的结果字典序号
            stage_data = result_list[i]
        
            #在这里对一个字典进行处理
        
            stage_data_dict[stage_name_dict['key']] = test(stage_data)
            stage_name_dict['key'] = stage_name   

    result = {}
    result['pipeline_stages_config_dict'] = stage_data_dict
    return result                              

def get_desktop(): 
    key = winreg.OpenKey(winreg.HKEY_CURRENT_USER,r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders')#利用系统的链表
    return winreg.QueryValueEx(key, "Desktop")[0] #返回的是Unicode类型数据

@app.route('/select', methods=('GET', 'POST'))
def select():

    result=[]  #是否只要每次的就可以 还是进行更新
    form = Form() #获取form
    s={} #再设s
    r=request.form  #r获取所有结果   
    for k,v in r.items():        
        s[k]=v
    result.append(s)
    
    # 获取数据  相当于read从select获取
    test_data = str(result)
    
    result_dict = DB_data()
    result_dict['pipeline_stages_config_dict'] = SE_data(ast.literal_eval(test_data[1:-1]))
    
    Desktop_path=str(get_desktop()) + '\\data_dict.txt'#Unicode转化为str
    pickle.dump(result_dict,open(Desktop_path,'wb'))

    return render_template('select.html', form=form)





@app.route('/<option_value>/')
def explanation(option_value):
    wb_path='excel.xlsx'  #数据表
    wb=openpyxl.load_workbook(wb_path) #打开        
    sheet=wb['Sheet1']  #打开sheet
    maxrow=sheet.max_row  #数据最大行
    ss=[] #设ss
    for i in range(1,maxrow+1): #循环所有数据
        d={} #设d
        value=sheet.cell(row=i,column=1).value #获取循环到的值
        value=value.strip()  #删除头尾的空格    
        if option_value==value: #如果找到了
            content=sheet.cell(row=i,column=2).value #获取其解释
            d['value']=value #形成将要输出的样子
            d['content']=content            
            ss.append(d) #添加到ss中
    return jsonify(ss) #https://www.jianshu.com/p/52cc29a0134d
