# 加载包 处理json和打开excel
import json
import openpyxl

# 打开测试数据json
with open('test.json', 'r') as f:
    data_list = json.load(f)  # json.load读书json数据
# print(data)

# 获取第一级选择字典
def s1option():    
    s = {} # 设选项内容字典
    for data in data_list: # 对读取的json数据循环处理 读取第一级
        s[data] = data  # 数据名与数据对应    
    # print(s)
    return s # 包含一级选择字典

# 获取第二级选择字典
def s2option():
    s = {} # 设选项内容字典
    for data in data_list: # 对读取的json数据循环处理 读取第一级
        d = [] # 第二级选项
        for k in data_list[data]:  # 对第一级数据循环，读取第二级            
            key = data + '-' + k  # 第一级-第二级 形成key
            d.append({key:k}) # 添加进第二级选项
        s[data] = d # 形成两级选项内容字典
    # print(s)
    return s

# 获取第三级选择字典
def s3option():
    s = {} #设选项内容字典
    for data in data_list: # 对读取的json数据循环处理 读取第一级
        for k in data_list[data]:  # 对第一级数据循环，读取第二级 
            key = data + '-' + k  # 第一级-第二级 形成key
            d = [] # 第三级选项
            for j in data_list[data][k]: # 循环获取第三级选项数据
                key2 = key + '-' + j # 重新组合 三级key               
                d.append({key2:j}) # 将数据向d中添加
            s[key] = d # 获得包含三级选择字典
    # print(s)           
    return s
            
# 获取第四级选择字典
def s4option():
    s = {} #设选项内容字典
    for data in data_list: # 对读取的json数据循环处理 读取第一级
        for k in data_list[data]:  # 对第一级数据循环，读取第二级 
            key = data + '-' + k  # 第一级-第二级 形成key        
            for j in data_list[data][k]: # 循环获取第三级选项数据
                key2 = key + '-' + j # 重新组合 三级key  
                d = [] # 第四级选项
                for f in data_list[data][k][j]: # 获取第四级选择key名
                    key3 = key2 + '-' + f #重新组合 四级key
                    d.append({key3:f})
                s[key2] = d #获得包含四级选择字典           
    # print(s)
    return s
            
# 输出到txt函数
def toTxt():
    
    # 四个函数
    s1 = s1option()
    s2 = s2option()
    s3 = s3option()
    s4 = s4option()
    
    # 打开选项txt,如果没有就新建一个
    f2 = open('option.txt','w',encoding='utf-8')
    
    # 向里写入数据
    f2.write('var s1option=['+str(s1)+'];'+'\n'+'var s2option=['+str(s2)+'];'+'\n'+'var s3option=['+str(s3)+'];'+'\n'+'var s4option=['+str(s4)+'];'+'\n')
    f2.close()


# 输出到excel表函数
def jsonToExcel():

    # 创建新的excel
    wb=openpyxl.Workbook()
    wb.create_sheet(index=0,title='Sheet1')           
    excel_name='excel.xlsx'
    wb.save(excel_name)

    # 写入excel
    wb=openpyxl.load_workbook(excel_name)        
    sheet=wb['Sheet1']

    r = 1 #行数
    for data in data_list: # 对读取的json数据循环处理 读取第一级
        for k in data_list[data]:  # 对第一级数据循环，读取第二级                        
            for j in data_list[data][k]: # 循环获取第三级选项数据                
                for f in data_list[data][k][j]: # 获取第四级选择key名
                    # print(f,data[i][k][j][f])
                    
                    key = data + '-' + k + '-' + j + '-' + f # 组合key值
                    sheet.cell(row=r, column=1).value = key # key值为第一列数据
                    sheet.cell(row=r, column=2).value = data_list[data][k][j][f] # 解释为第二列数据
                    r += 1 #行数递增
    # 储存
    wb.save(excel_name)
                
# 主函数
def main():
    # 执行生成txt excel表的函数
    toTxt()
    jsonToExcel()

# 启动
if __name__ == '__main__':
    main()
